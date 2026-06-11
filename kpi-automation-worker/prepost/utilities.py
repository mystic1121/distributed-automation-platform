# -*- encoding: utf-8 -*-

import os
import yaml
import re

import numpy as np
import pandas as pd

from copy import deepcopy
from tqdm import tqdm as TQ
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter, column_index_from_string


def idx_to_letter(idx):
    """Convert Excel column index to letter (1-based)."""
    result = ''
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ─── openpyxl helpers ───────────────────────────────────────────────────────

def _clean_val(v):
    """Convert numpy nan/inf and None to None so openpyxl won't crash."""
    if v is None:
        return None
    try:
        if np.isnan(v) or np.isinf(v):
            return None
    except (TypeError, ValueError):
        pass
    # unwrap numpy scalars to Python native types
    if hasattr(v, 'item'):
        return v.item()
    return v


def _get_last_row_in_col(ws, col):
    """Return the last row index (1-based) that has a non-None value in *col*."""
    for r in range(ws.max_row, 0, -1):
        if ws.cell(r, col).value is not None:
            return r
    return 0


def _write_row(ws, row, start_col, data):
    """Write an iterable as a single row starting at (row, start_col)."""
    for i, val in enumerate(data):
        ws.cell(row=row, column=start_col + i, value=_clean_val(val))


def _write_column(ws, start_row, col, data):
    """Write an iterable as a single column starting at (start_row, col)."""
    for i, val in enumerate(data):
        ws.cell(row=start_row + i, column=col, value=_clean_val(val))


def _write_df_no_header_no_index(ws, start_row, start_col, df):
    """Write DataFrame values only — no column headers, no index.

    Equivalent to xlwings:  ws[cell].options(header=False, index=False).value = df
    """
    for ri, row_data in enumerate(df.values):
        for ci, val in enumerate(row_data):
            ws.cell(row=start_row + ri, column=start_col + ci, value=_clean_val(val))


def _write_df_with_header_and_index(ws, start_row, start_col, df):
    """Write DataFrame with column headers and index columns.

    Equivalent to xlwings default:  ws[cell].value = df
    Handles single-level and multi-level index.
    """
    idx_levels = df.index.nlevels

    # header row — index name(s) then column names
    if isinstance(df.index, pd.MultiIndex):
        for li, name in enumerate(df.index.names):
            ws.cell(row=start_row, column=start_col + li, value=name)
    else:
        ws.cell(row=start_row, column=start_col, value=df.index.name)
    for ci, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + idx_levels + ci, value=str(col_name))

    # data rows
    for ri, (idx_val, row_data) in enumerate(df.iterrows()):
        r = start_row + 1 + ri
        if isinstance(df.index, pd.MultiIndex):
            for li, iv in enumerate(idx_val):
                ws.cell(row=r, column=start_col + li, value=_clean_val(iv))
        else:
            ws.cell(row=r, column=start_col, value=_clean_val(idx_val))
        for ci, val in enumerate(row_data):
            ws.cell(row=r, column=start_col + idx_levels + ci, value=_clean_val(val))


def _write_df_with_header_no_index(ws, start_row, start_col, df):
    """Write DataFrame with column headers but without index.

    Equivalent to xlwings:  ws[cell].options(index=False).value = df
    """
    for ci, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + ci, value=str(col_name))
    for ri, row_data in enumerate(df.values):
        for ci, val in enumerate(row_data):
            ws.cell(row=start_row + 1 + ri, column=start_col + ci, value=_clean_val(val))


# ─── excel cleanup ──────────────────────────────────────────────────────────

def perform_excel_cleanup(wb_or_path):
    """Post-process the 'aggregated' sheet — find a junk column, insert a clean
    column H, propagate formulas/values, apply KPI-specific divisors.

    Accepts either a file path (str) or an already-open openpyxl Workbook object.
    When given a workbook object the caller is responsible for saving.
    """
    is_path = isinstance(wb_or_path, str)

    if is_path:
        print(f"\n{'='*60}")
        print(f" STARTING: {os.path.basename(wb_or_path)}")
        print(f"{'='*60}")
        wb = load_workbook(wb_or_path)
    else:
        wb = wb_or_path
        print(f" STARTING CLEANUP: {getattr(wb, '_archive', None) and 'workbook' or 'workbook'}")

    try:
        if 'aggregated' not in [s.title for s in wb.worksheets]:
            print(" Sheet 'aggregated' not found. Skipping cleanup.")
            return

        ws = wb['aggregated']

        last_row = _get_last_row_in_col(ws, 2)   # column B
        last_col = ws.max_column
        scan_limit = min(last_col, 50)

        print(f" Range Detected: Row 1 to {last_row}")

        # ── STEP 1: IDENTIFY JUNK COLUMN ──────────────────────────────────
        target_junk_idx = None
        junk_set = {1000, "1000", 1000.0, 0, 0.0, "0", "0.0", "0.00", None, "", " "}
        DENSITY_THRESHOLD = 0.40

        for col in range(8, scan_limit + 1):
            values = [ws.cell(r, col).value for r in range(3, last_row + 1)]
            junk_count = sum(
                1 for v in values
                if v in junk_set or str(v).strip() in {"0", "1000", ""}
            )
            density = junk_count / len(values) if values else 0
            if density >= DENSITY_THRESHOLD:
                target_junk_idx = col
                print(f" JUNK DETECTED: Column {idx_to_letter(col)} (Density: {density:.2%})")
                break

        if not target_junk_idx:
            print(" No junk pattern found.")
            return

        # ── STEP 2: INSERT NEW COLUMN H ────────────────────────────────────
        print(" Inserting new column at H...")
        ws.insert_cols(8)
        shifted_junk_idx = target_junk_idx + 1

        # ── STEP 3: MOVE FORMULA/VALUE INTO NEW H (SAFE) ──────────────────
        print(" Moving formulas/values into new H (with safe propagation)...")
        special_kpis = {
            "session setup success rate",
            "overall_5g_inter frequency handover success rate (%)"
        }
        exclude_vals = {
            1000, "1000", 1000.0,
            100, "100", 100.0,
            1000000, "1000000", 1000000.0,
            None, "", " "
        }

        for r in range(1, last_row + 1):
            kpi_name = str(ws.cell(r, 2).value).strip().lower()
            raw = ws.cell(r, shifted_junk_idx).value

            is_formula = isinstance(raw, str) and raw.startswith("=")
            formula_source = raw if is_formula else ""
            val_source = None if is_formula else raw

            if val_source in exclude_vals and not is_formula:
                continue

            start_col = 8
            end_col = shifted_junk_idx - 1

            for c in range(start_col, end_col + 1):
                cell_target = ws.cell(r, c)
                current_col_letter = idx_to_letter(c)

                if is_formula:
                    if kpi_name in special_kpis:
                        new_formula = re.sub(
                            r"([A-Z]+)(\d+)", rf"{current_col_letter}\2", formula_source
                        )
                    else:
                        col_offset = c - start_col
                        base_letter = idx_to_letter(4 + col_offset)
                        new_formula = re.sub(
                            r"!([A-Z]+)(\d+)", rf"!{base_letter}\2", formula_source
                        )
                    cell_target.value = new_formula
                    cell_target.number_format = "0.00"
                else:
                    if val_source not in exclude_vals:
                        cell_target.value = val_source
                        cell_target.number_format = "0.00"

        # ── STEP 4: HORIZONTAL PROPAGATION FOR SPECIAL KPIs ───────────────
        print(f" Applying formulas across columns up to {idx_to_letter(shifted_junk_idx - 1)} safely...")
        for r in range(1, last_row + 1):
            base_val = ws.cell(r, 8).value
            if not (isinstance(base_val, str) and base_val.startswith("=")):
                continue
            base_formula = base_val
            try:
                for c in range(8, shifted_junk_idx):
                    cell_h_val = ws.cell(r, 8).value
                    if cell_h_val in exclude_vals:
                        continue
                    current_col_letter = idx_to_letter(c)
                    kpi_name = str(ws.cell(r, 2).value).strip().lower()
                    if kpi_name in special_kpis:
                        new_formula = re.sub(
                            r"([A-Z]+)(\d+)", rf"{current_col_letter}\2", base_formula
                        )
                    else:
                        col_offset = c - 8
                        base_letter = idx_to_letter(4 + col_offset)
                        new_formula = re.sub(
                            r"!([A-Z]+)(\d+)", rf"!{base_letter}\2", base_formula
                        )
                    ws.cell(r, c).value = new_formula
                    ws.cell(r, c).number_format = "0.00"
            except Exception as e:
                print(f" Warning: Row {r} failed propagation: {e}")

        # ── STEP 5: DIVIDE BY KPI-SPECIFIC DIVISORS ───────────────────────
        print("Dividing numeric data by KPI-specific divisors...")

        divide_kpi_keywords1 = ["userthroughput_dl"]           # ÷ 1 000
        divide_kpi_keywords2 = ["5g_drb_pdcppacketlossrateul_5qi1"]  # ÷ 1 000 000
        divide_kpi_keywords3 = ["5g_dlpacketdelay-5qi1(packetdelay)_ran"]  # ÷ 100

        for r in range(1, last_row + 1):
            kpi_name_raw = ws.cell(r, 2).value
            if not kpi_name_raw:
                continue
            kpi_name = str(kpi_name_raw).lower().replace(" ", "").replace("\xa0", "")

            for c in range(8, shifted_junk_idx):
                cell = ws.cell(r, c)
                val = cell.value
                if not isinstance(val, (int, float)) or val == 0:
                    continue

                if any(kw in kpi_name for kw in divide_kpi_keywords1):
                    cell.value = val / 1000
                    cell.number_format = "0.00"
                elif any(kw in kpi_name for kw in divide_kpi_keywords2):
                    cell.value = val / 1000000
                    cell.number_format = "General"
                elif any(kw in kpi_name for kw in divide_kpi_keywords3):
                    cell.value = val / 100
                    cell.number_format = "0.00"

        # ── FINALISE ──────────────────────────────────────────────────────
        print(" Finalizing cleanup...")
        if is_path:
            wb.save(wb_or_path)
            wb.close()
            print(f" SUCCESS: Processed standalone {os.path.basename(wb_or_path)}\n")
        else:
            print(f" SUCCESS: Cleanup done on in-memory workbook")

    except Exception as e:
        print(f" ERROR in cleanup: {e}")
        import traceback
        traceback.print_exc()


def read_file(filepath: str, date_format: str = "%d/%m/%Y %H:%M",
              fillna: float = -np.inf, JCPFormat: bool = False) -> pd.DataFrame:
    try:
        data = pd.read_csv(filepath, encoding='utf-8', sep=None, engine='python')
    except UnicodeDecodeError:
        print(f"Warning: UTF-8 failed for {filepath}. Trying cp1252...")
        try:
            data = pd.read_csv(filepath, encoding='cp1252', sep=None, engine='python')
        except UnicodeDecodeError:
            print(f"Warning: cp1252 also failed. Trying latin-1...")
            data = pd.read_csv(filepath, encoding='latin-1', sep=None, engine='python')

    data["datetime"] = data[["Date", "Time"]].apply(lambda x: f"{x[0]} {x[1]}", axis=1)
    data["datetime"] = pd.to_datetime(data["datetime"], format=date_format)
    data = data.drop(columns=["Date", "Time"]).set_index(["datetime"]).reset_index()

    if JCPFormat:
        data.fillna(fillna, inplace=True)
    else:
        for col in TQ(data.columns, desc="type casting"):
            try:
                data[col] = data[col].apply(
                    lambda x: fillna if (x == "-") or (x == "NaN") else float(x)
                )
            except ValueError:
                pass
            except TypeError:
                pass
        data.dropna(axis=1, inplace=True)

    return data.copy()


# ─── write_excel ────────────────────────────────────────────────────────────

def write_excel(
    template_filepath: str, ROOT: str, PROJECT_CODE: str,
    datetime_reference: dict, LOCATION_NAME: str,
    reports: object,
    REPORT_DATE, REPORT_TYPE,
    OUTPUT_DIR, OUTPUT_FILE,
    aggregated_sheet_remarks,
    aggregated_sheet_frame,
    aggregated_daily_sheet_frame,
    post_process: bool = False,
    app=None          # kept for signature compatibility — ignored
) -> bool:
    """Open *template_filepath*, populate all report sheets, save to OUTPUT_DIR/OUTPUT_FILE."""
    try:
        wb = load_workbook(template_filepath)

        # ── about sheet ───────────────────────────────────────────────────
        with open(os.path.join(ROOT, "template", PROJECT_CODE, "meta.yaml")) as f:
            meta = yaml.load(f, Loader=yaml.FullLoader)

        about_ws = wb["about"]
        for cell, value in meta["about"].items():
            if cell == "J7":
                value = eval(value)
            about_ws[cell] = value
        about_ws["L12"] = f"Report Date: {REPORT_DATE}\nReport Type: {REPORT_TYPE}"

        # ── datetime arrays ───────────────────────────────────────────────
        # each key is "DD-MM-YYYY HH:MM"  →  split into [date, time]
        datetimes = np.array([dt_str.split() for dt_str in datetime_reference.keys()])

        # ── Summary Sheet ─────────────────────────────────────────────────
        summary_ws = wb["Summary Sheet"]
        summary_ws["A1"] = f"KPI Report for {LOCATION_NAME}"
        _write_row(summary_ws, 1, 5, datetimes[:, 0])           # E1
        if REPORT_TYPE != "daily":
            _write_row(summary_ws, 2, 5, datetimes[:, 1])       # E2
        _write_df_no_header_no_index(summary_ws, 3, 5, reports[0])  # E3

        # ── aggregated sheet ──────────────────────────────────────────────
        agg_ws = wb["aggregated"]
        _write_column(agg_ws, 3, 7, [_clean_val(v) for v in aggregated_sheet_remarks])  # G3
        _write_df_no_header_no_index(agg_ws, 3, 8, aggregated_sheet_frame)               # H3
        _write_row(agg_ws, 1, 8, datetimes[:, 0])               # H1
        if REPORT_TYPE != "daily":
            _write_row(agg_ws, 2, 8, datetimes[:, 1])           # H2

        # ── Nodewise Report ───────────────────────────────────────────────
        nodewise_ws = wb["Nodewise Report"]
        _write_df_with_header_and_index(nodewise_ws, 1, 1, reports[1])

        # ── Error Remarks ─────────────────────────────────────────────────
        error_ws = wb["Error Remarks"]
        _write_df_with_header_and_index(error_ws, 1, 1, reports[2])

        # ── aaggregated-daily ─────────────────────────────────────────────
        agg_daily_ws = wb["aaggregated-daily"]
        _write_df_with_header_no_index(agg_daily_ws, 1, 3, aggregated_daily_sheet_frame)  # C1

        # ── optional post-processing ──────────────────────────────────────
        if post_process:
            perform_excel_cleanup(wb)

        wb.save(os.path.join(OUTPUT_DIR, OUTPUT_FILE))
        wb.close()
        return True

    except Exception as e:
        print(f" ERROR inside write_excel: {e}")
        import traceback
        traceback.print_exc()
        return False
